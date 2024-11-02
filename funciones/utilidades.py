import os
import openpyxl
import datetime as dt
import time


def guardar_estadísticas(resultado:str, intentos_usados:int, max_intentos:int, max_rango:int, puntos:int, dificultad:str, modo_juego:str):
    '''Guarda las estadisticas de juego en un archivo Excel.
    Se guardarán los datos requeridos en los argumentos:
    - Resultado (str): si el jugador adivinador es "ganador" o "perdedor".
    - Intentos_usados (int): intentos empleados por el jugador adivinador.
    - Max_intentos (int): número máximo de intentos permitidos.
    - Max_rango (int): número maximo del rango del número objetivo.

    Los datos se almacenarán junto con el nombre del jugador adivinador, la fecha y hora de la partida.
    '''
    print("Por favor Introduce tu nombre para almacenar las estadísticas de juego: ")
    try:
        nombre = input()
    except Exception as e:
        print("Error al introducir el nombre. Se guardará como 'Desconocido'")
        nombre = "Desconocido"    
    
    # Nombre de las columnas a guardar
    columnas = ["Nombre", "Timestamp", "Modo de Juego", "Resultado", "Intentos Usados", "Max Intentos", "Max Rango", "Puntos", "Dificultad"]

    # Comprobar si existe el archivo de estadísticas
    if os.path.exists("estadisticas.xlsx"):

        # si existe, cargamos el archivo.
        try:
            archivo_excel = openpyxl.load_workbook("estadisticas.xlsx")
        
        # El archifvo podría estar en uso, esperamos 5 segundos y volvemos a intentar.
        # En simulación automatizada podría estar cerrandose el archivo.
        except PermissionError:
            time.sleep(5)
            archivo_excel = openpyxl.load_workbook("estadisticas.xlsx")

        # Comprobamos si el archivo contiene una pestaña llamada estadísticas
        if not "estadísticas" in archivo_excel.sheetnames:
            # Creramos la pestaña estadisticas y escribimos los nombres de las columnas.
            archivo_excel.create_sheet("estadísticas")
            pestaña = archivo_excel["estadísticas"]
            pestaña.append(columnas)
        else:
            # Si la pestaña existe la seleccionamos.
            pestaña = archivo_excel["estadísticas"]

    else:
        # Creamos el archivos y la pestaña estadísticas.
        archivo_excel = openpyxl.Workbook()
        archivo_excel.create_sheet("estadísticas")

        # Seleccionamos la pestaña estadísticas.
        pestaña = archivo_excel["estadísticas"]
        archivo_excel.active = pestaña

        # Escribimos los nombres de las columnas.
        pestaña.append(columnas)

        # Borrar pestañas innecesarias
        lista_pestañas = archivo_excel.sheetnames
        lista_pestañas.remove("estadísticas")
        for pestaña_listada in lista_pestañas:
            archivo_excel.remove(worksheet=archivo_excel[pestaña_listada])

    # Escribimos los datos al final de los datos existentes y guardamos el archivo.
    pestaña.append([nombre, dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), modo_juego, resultado, intentos_usados, max_intentos, max_rango, puntos, dificultad])
    try:
        archivo_excel.save("estadisticas.xlsx")
    except PermissionError:
        time.sleep(5)
        archivo_excel.save("estadisticas.xlsx")

    print("Estadísticas guardadas con éxito.\n")


def validar_selección(selección, opción_max:int, opción_min:int=1) -> int:
    ''' Validación de la entrada de un número entero con caráteres numéricos dentro del rango determinado.
    Parámetros:
    - selección (int): número introducido por el usuario.
    - opción_max (int): número máximo del rango de opciones.
    - opción_min (int): número mínimo del rango de opciones. Por defecto es 1.

    Retorna:
    Si la validaición es correcta, retorna el valor seleccionado. Si la validación no es correcta, retorna 0.
    '''
    try:
        selección = int(selección)
    except ValueError:
        print("Por favor, introduce un número entero válido.\n")
        return 0
    else:
        if selección not in range(opción_min, opción_max + 1):
            print(f"Introduce un numero entre {opción_min} y {opción_max}.\n")
            return 0
        else:
            return selección
        

def gestión_menu(opciones:dict, msg_intro:str="Selecciona una opción: ", msg_accion:str="¿Que opción eliges?: ") -> int:
    '''Gestión de un menú de opciones.
    Parámetros:
    - opciones (dict): diccionario con las opciones del menú. Las jeys de diccionario debe sen números enteros correspondientes a cada una de las opciones.

    Retorna:
    La opción seleccionada por el usuario.
    '''
    selección = 0
    while selección not in opciones.keys():
        print(f"\n{msg_intro}")
        for key, value in opciones.items():
            print(f"{key}. {value}")
        print(f"\n{msg_accion}", end="")
        selección = validar_selección(input(), opción_max=len(opciones))

    return selección

if __name__ == "__main__":
    datos = {
        "resultado": "Ganador",
        "intentos_usados": 5,
        "max_intentos": 20,
        "max_rango": 1000,
        "puntos": 100,
        "dificultad": "Fácil",
        "modo_juego": "Solitario"
    }
    guardar_estadísticas(**datos)